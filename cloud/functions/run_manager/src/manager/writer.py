from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
import pandas as pd


def write_cumulative_result(frames: dict, save_path: str):
    """ writes out summary frames received for all results to Master Results.xlsx """
    workbook = Workbook()
    workbook.remove(workbook.active)

    sheets = {}
    for name, (df, currency_cols) in frames.items():
        sheets[name] = workbook.create_sheet(title=name)
        write_dataframe(sheets[name], df)
        sheets[name].freeze_panes = 'A2'  # A2 appears to freeze top row, B2 first column.
        columns_to_currency(sheets[name], currency_cols)

    tidy_workbook(workbook)
    workbook.save(save_path)


def write_result(folder_path: str, file_name: str, result: pd.DataFrame, draw_result: dict, prize_breakdown: dict):
    """ writes out result of a given draw. This includes the:
        * draw result - how many matches achieved,
        * draw outcome - balls drawn,
        * prize breakdown - prize breakdown for the draw.
    """
    workbook = Workbook()
    workbook.remove(workbook.active)  # remove default 'Sheet'

    result_sheet = workbook.create_sheet('Result', 0)
    draw_sheet = workbook.create_sheet('Draw Outcome', 1)
    prize_breakdown_sheet = workbook.create_sheet('Prize Breakdown', 2)

    write_dataframe(result_sheet, result)

    write_dict(draw_sheet, draw_result, [])
    write_dict(prize_breakdown_sheet, prize_breakdown, ['Match Type', 'Prize Per UK Winner'])

    tidy_workbook(workbook)
    workbook.save(f'{folder_path}/{file_name}.xlsx')


def write_dataframe(sheet: Worksheet, df: pd.DataFrame, index=False, header=True):
    for row in dataframe_to_rows(df, index=index, header=header):
        sheet.append(row)


def write_dict(sheet: Worksheet, data: dict, col_names: list):
    if col_names:
        sheet.append(col_names)
    for item in data.items():
        sheet.append(item)


def tidy_workbook(workbook: Workbook):
    """ Iterates of all sheets in a workbook, all used columns in a sheet, all used cells in a column and formats
        them by: aligning values and ensuring column with is wide enough for given values."""
    desired_alignment = Alignment(horizontal="center", vertical="center")
    for name in workbook.sheetnames:
        sheet = workbook[name]
        for column_cells in sheet.columns:
            desired_length = sheet.column_dimensions[column_cells[0].column_letter].width
            for cell in column_cells:
                desired_length = max(len(as_text(cell.value)), desired_length)
                cell.alignment = desired_alignment
            sheet.column_dimensions[column_cells[0].column_letter].width = desired_length


def as_text(value):
    if value is None:
        return ""
    else:
        return str(value)


def columns_to_currency(sheet: Worksheet, columns: list):
    """ iterates over all cells in selected columns to convert their format to currency"""
    for column_cells in sheet.columns:
        if column_cells[0].value in columns:
            for cell in column_cells[1:]:
                # see https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/numbers.html for formats
                cell.number_format = '"£"#,##0.00_-'
