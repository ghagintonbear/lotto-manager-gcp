from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, NamedStyle
import pandas as pd


def write_cumulative_result(cumulative: pd.DataFrame, aggregated: pd.DataFrame):
    workbook = Workbook()
    workbook.remove(workbook.active)

    all_results_sheet = workbook.create_sheet('All Results')
    aggregated_sheet = workbook.create_sheet('Aggregated')

    write_dataframe(all_results_sheet, cumulative)
    write_dataframe(aggregated_sheet, aggregated)

    columns_to_currency(all_results_sheet, ['Winnings', 'Winning per Person', 'Cumulative Winnings per Person'])
    columns_to_currency(aggregated_sheet, ['Winnings', 'Winning per Person'])

    all_results_sheet.freeze_panes = 'A2'  # A2 appears to freeze top row, B2 first column.

    tidy_workbook(workbook)
    workbook.add_named_style(NamedStyle(name='highlight'))
    workbook.save('./Master Results.xlsx')


def write_result(folder_path: str, file_name: str, result: pd.DataFrame, draw_result: dict, prize_breakdown: dict):
    workbook = Workbook()
    workbook.remove(workbook.active)  # remove default 'Sheet'

    result_sheet = workbook.create_sheet('Result', 0)
    draw_sheet = workbook.create_sheet('Draw Outcome', 1)
    prize_breakdown_sheet = workbook.create_sheet('Prize Breakdown', 2)

    write_dataframe(result_sheet, result)

    write_dict(draw_sheet, draw_result, [])
    write_dict(prize_breakdown_sheet, prize_breakdown, ['Match Type', 'Prize Per UK Winner'])

    tidy_workbook(workbook)
    workbook.save(f'{folder_path}/{file_name}.xlsx')


def write_dataframe(sheet: Worksheet, df: pd.DataFrame, index=False, header=True):
    for row in dataframe_to_rows(df, index=index, header=header):
        sheet.append(row)


def write_dict(sheet: Worksheet, data: dict, col_names: list):
    if col_names:
        sheet.append(col_names)
    for item in data.items():
        sheet.append(item)


def tidy_workbook(workbook: Workbook):
    desired_alignment = Alignment(horizontal="center", vertical="center")
    for name in workbook.sheetnames:
        sheet = workbook.get_sheet_by_name(name)
        for column_cells in sheet.columns:
            desired_length = sheet.column_dimensions[column_cells[0].column_letter].width
            for cell in column_cells:
                desired_length = max(len(as_text(cell.value)), desired_length)
                cell.alignment = desired_alignment
            sheet.column_dimensions[column_cells[0].column_letter].width = desired_length


def as_text(value):
    if value is None:
        return ""
    else:
        return str(value)


def columns_to_currency(sheet: Worksheet, columns: list):
    for column_cells in sheet.columns:
        if column_cells[0].value in columns:
            for cell in column_cells[1:]:
                # https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/numbers.html
                cell.number_format = '"£"#,##0.00_-'
